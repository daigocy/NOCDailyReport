#! python2.7
# -*- encoding:utf-8 -*-
import os
import urllib2
import cookielib
import urllib
import re
import time
import datetime
import _winreg
import warnings
import csv

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


warnings.filterwarnings('ignore')
today = datetime.datetime.now()
yesterday = today + datetime.timedelta(days=-1)
yesterday_str = str(yesterday.strftime("%Y-%m-%d"))
query_day_list = []
print u"请输入日期，如 2017-01-01"
print u"或者直接回车使用昨天日期" + yesterday_str
while len(query_day_list) == 0:
    input_day = raw_input(":")
    if input_day == '':
        query_day_list = [yesterday_str]
        print u'使用默认日期:' + yesterday_str
    else:
        query_day_list = re.findall(r'(2[0-9]{3}-[0-1][0-9]-[0-3][0-9])', input_day)
        # 判断日期格式
        if len(query_day_list) == 0:
            print u'格式错误'
        else:
            try:
                date_check = datetime.datetime.strptime(query_day_list[0], "%Y-%m-%d")
                # 判断日期合法
            except Exception as e:
                print u'日期错误'
                query_day_list = []
query_day = query_day_list[0]
start_end_times = [
        [query_day+" 00:00:00", query_day+" 18:59:59"],
        [query_day+" 19:00:00", query_day+" 23:59:59"]
    ]
# query_set = [
#     u'所有操作慢',
#     u'浏览网页慢',
#     u'他网视频慢',
#     u'下载慢',
#     u'游戏慢',
#     u'游戏问题'
# ]
query_codes = ['3222', '3223', '3224', '3225', '3226', '3229']
# firedebug得出相应查询码
postUrl = 'http://218.108.129.189:19090/ccps/login.action'
cookie = cookielib.CookieJar()
handler = urllib2.HTTPCookieProcessor(cookie)
post_opener = urllib2.build_opener(handler)
# 将cookies绑定到一个opener    cookie由cookielib自动管理
username = 'monitor'
password = 'Wasu@2016'
postData = {'staff.password': password,
            'staff.wcode': username,
            }
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3',
    'Host': '218.108.129.189:19090',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:52.0) Gecko/20100101 Firefox/52.0'
}
data = urllib.urlencode(postData)
request1 = urllib2.Request(postUrl, data, headers)
response1 = post_opener.open(request1)
# 完成IPCC登录
original_list = []
# IPCC 故障量表
for time_list in start_end_times:
    for query_code in query_codes:
        get_parameter = {
            'activity.activityId': '',
            'activity.activitySource': '',
            'activity.activityType': '',
            'activity.blank5': query_code,
            'activity.blank6': '',
            'activity.busSubclass': query_code,
            'activity.busType': '',
            'faultWorkflow': '',
            'activity.cityName': '杭州市区',
            'activity.custAddr': '',
            'activity.custName': '',
            'activity.custTel': '',
            'activity.depIds': '27,28,29,30',
            'activity.depLCs': '网管中心',
            'activity.depNames': '96171,青鸟队,鸿雁队,百灵队',
            'activity.disState': '',
            'activity.productClass': '51',
            'activity.productId': '',
            'activity.refOrderId': '',
            'activity.regionId': '1051',
            'activity.remark': '',
            'activity.reqSource': '',
            'activity.specialNeeds': '',
            'activity.tempEndTime': time_list[1],
            'activity.tempStartTime': time_list[0],
            'activity.wcode': ''
        }
        get_url = urllib.urlencode(get_parameter)
        host_url = 'http://218.108.129.189:19090/ccps/activity/findActivitiesListForView.action'
        full_url = host_url + '?' + get_url
        response = post_opener.open(full_url)
        soup = BeautifulSoup(response, "html.parser")
        summary = soup.find("span", class_=['style2']).get_text()
        # 利用beautifulsoup抓取目标tag内的文本
        reg = u'共(\d+)条记录'
        num = re.findall(reg, summary, re.MULTILINE)
        original_list.append(int(num[0]))
        # 将每类查询的汇总条数加入 original_list

wb = Workbook()
ws = wb.active
font1 = Font(size=9, )
align = Alignment(horizontal='center')
num_format = '0.00%'
side = Side(style='thin', color="000000")
border = Border(left=side, right=side, top=side, bottom=side)

# 新建excel类，激活sheet
sum1 = 0
sum2 = 0
# 合计
line = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
# excel列名
query_set = [
    u'所有操作慢',
    u'浏览网页慢',
    u'他网视频慢',
    u'下载慢',
    u'游戏慢',
    u'游戏问题'
]
for i in range(0, 6):
    ws.cell(line[i + 2] + '1').value = query_set[i]
    ws.cell(line[i + 2] + '2').value = original_list[i]
    sum1 = sum1 + original_list[i]
for i in range(6, 12):
    ws.cell(line[i + 3] + '1').value = query_set[i - 6]
    ws.cell(line[i + 3] + '2').value = original_list[i]
    sum2 = sum2 + original_list[i]
sum3 = sum1 + sum2
ws.cell("A1").value = u'总数'
ws.cell("B1").value = u'合计'
ws.cell("I1").value = u'合计'
ws.cell("A2").value = sum3
ws.cell("B2").value = sum1
ws.cell("I2").value = sum2
for cell in ws["A2:O2"][0]:
    cell.font = font1
    cell.alignment = align
    cell.border = border

postUrl = 'http://172.25.254.48/index.php'
username = 'reporter'
password = 'reporter'
postData = {
    'action': 'login',
    'login_password': password,
    'login_username': username,
            }
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3',
    'Connection': 'keep-alive',
    'Host': '172.25.254.47',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:52.0) Gecko/20100101 Firefox/52.0'
}
data = urllib.urlencode(postData)
request0 = urllib2.Request('http://172.25.254.48/')
response0 = post_opener.open(request0)
request2 = urllib2.Request(postUrl, data, headers)
response2 = post_opener.open(request2)
# 登录CartiEZ成功
start_time = time.strptime(query_day, "%Y-%m-%d")
start_time_float = time.mktime(start_time)
star_time_int = int(start_time_float)
end_time_int = star_time_int + (24*60*60-60)
start_time_str = str(star_time_int)
end_time_str = str(end_time_int)
graph_codes = ['93', '95', '2270', '2997']
query_set2 = [
    u'ISP8-1延时', u'ISP8-1丢包', u'ISP1-2延时', u'ISP1-丢包2', u'ISP1-3延时', u'ISP1-3丢包',
    u'GD10G延时', u'GD10G丢包', u'SZ10G延时', u'SZ10G丢包', u'华通', u'杭网IDC'
]
for i in range(0, 12):
    ws.cell(row=4, column=i+1, value=query_set2[i])

host_url = 'http://172.25.254.48/graph_xport.php'


def write_bgp1(read_csv):
    read_list = list(read_csv)
    required_col = [4, 2, 3, 5, 6]
    results = [0] * 5
    for each_line in read_list[10:]:
        for i in range(0, 5):
            if len(each_line):
                x = float(each_line[required_col[i]])
                if x > results[i]:
                    results[i] = x
    for i in range(0, 5):
        results[i] = round(results[i], 2)
        cell = ws.cell(row=5, column=i * 2 + 1)
        cell.value = results[i]
        cell.font = font1
        cell.alignment = align
        cell.border = border


def write_bgp2(read_csv):
    read_list = list(read_csv)
    required_col = [4, 2, 3, 5, 6]
    results = [0] * 5
    for each_line in read_list[10:]:
        for i in range(0, 5):
            if len(each_line):
                x = float(each_line[required_col[i]])
                if x > results[i]:
                    results[i] = x
    for i in range(0, 5):
        results[i] = round(results[i]/100, 4)
        cell = ws.cell(row=5, column=i * 2 + 2)
        cell.value = results[i]
        cell.font = font1
        cell.alignment = align
        cell.number_format = num_format
        cell.border = border


def write_hw(read_csv):
    read_list = list(read_csv)
    required_col = 5
    result = 0
    for each_line in read_list[10:]:
        if len(each_line):
            x = float(each_line[required_col])
            if x > result:
                result = x
    result = round(result / 1000000000, 2)
    result = int(result * 1000)
    cell = ws.cell(row=5, column=11, value=result)
    cell.font = font1
    cell.alignment = align
    cell.border = border


def write_ht(read_csv):
    read_list = list(read_csv)
    required_col = 6
    result = 0
    for each_line in read_list[10:]:
        if len(each_line):
            x = float(each_line[required_col])
            if x > result:
                result = x
    result = round(result/1000000000, 2)
    cell = ws.cell(row=5, column=12, value=result)
    cell.font = font1
    cell.alignment = align
    cell.border = border


write_methods = [write_bgp1, write_bgp2, write_hw, write_ht]
for i in range(0, 4):
    get_param = {
        'graph_end': end_time_str,
        'graph_start': start_time_str,
        'local_graph_id': graph_codes[i],
        'rra_id': '0',
        'view_type': 'tree'
    }
    get_url = urllib.urlencode(get_param)
    full_url = host_url + '?' + get_url
    response = post_opener.open(full_url)
    csv_file = csv.reader(response)
    write_methods[i](csv_file)

reg_key = _winreg.OpenKey(_winreg.HKEY_CURRENT_USER,
                          r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders',)
desk_path = _winreg.QueryValueEx(reg_key, "Desktop")[0]
# 查询注册表 获取桌面路径
dir_path = desk_path + r'\日报'.decode('utf-8') + query_day
if not os.path.exists(dir_path):
    os.makedirs(dir_path)
try:
    wb.save(filename=dir_path + r'\日报'.decode('utf-8') + query_day + '.xlsx')
    print u'生成日报' + query_day + '.xlsx'
except Exception as e:
    print str(e)

pic_names = [r'\延时.png', r'\丢包.png', r'\华通内容中心.png', r'\杭网内容中心.png']
host_url = 'http://172.25.254.48/graph_image.php'
for i in range(0, 4):
    get_param = {
        'graph_end': end_time_str,
        'graph_start': start_time_str,
        'local_graph_id': graph_codes[i],
        'rra_id': '0',
        'view_type': 'tree'
    }
    get_url = urllib.urlencode(get_param)
    full_url = host_url + '?' + get_url
    response = post_opener.open(full_url)
    pic = response.read()
    f = open(dir_path + pic_names[i].decode('utf-8'), 'wb')
    f.writelines(pic)
    f.close()
    print '生成'.decode('utf-8') + dir_path + pic_names[i].decode('utf-8')
time.sleep(5)
